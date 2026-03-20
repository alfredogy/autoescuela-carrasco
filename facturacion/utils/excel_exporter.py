from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def exportar_facturas_excel(anio, autoescuela, trimestre=None):
    from facturacion.models import Factura

    qs = Factura.objects.filter(anio=anio, autoescuela=autoescuela)
    if trimestre:
        qs = qs.filter(trimestre=trimestre)
    qs = qs.order_by('numero_factura')

    wb = Workbook()
    ws = wb.active
    ws.title = f'Facturas {anio}' if not trimestre else f'Trimestre {trimestre}'

    headers = ['CURSO', 'Nº FACTURA', 'FECHA', 'NOMBRE Y APELLIDOS', 'DNI',
               'BASE IMPONIBLE', 'IVA', 'TASAS', 'TOTAL',
               'DIRECCION', 'CP', 'MUNICIPIO', 'PROVINCIA']
    ws.append(headers)

    for f in qs:
        ws.append([
            f.curso,
            f.numero_factura,
            f.fecha.strftime('%d/%m/%Y') if f.fecha else '',
            f.nombre_factura,
            f.dni_factura,
            float(f.base_imponible),
            float(f.iva),
            float(f.tasas),
            float(f.total),
            f.direccion_factura,
            f.cp_factura,
            f.municipio_factura,
            f.provincia_factura,
        ])

    for row in range(2, ws.max_row + 1):
        for col in (6, 7, 8, 9):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.number_format = '#,##0.00'

    col_widths = [8, 15, 12, 30, 12, 15, 12, 12, 12, 30, 8, 15, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def exportar_alumnos(autoescuela):
    from facturacion.models import Alumno

    wb = Workbook()
    ws = wb.active
    ws.title = 'Alumnos'

    headers = ['NOMBRE Y APELLIDOS', 'DNI', 'DIRECCION', 'CP', 'MUNICIPIO', 'PROVINCIA']
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)

    for a in Alumno.objects.filter(autoescuela=autoescuela).order_by('nombre'):
        ws.append([a.nombre, a.dni, a.direccion, a.codigo_postal, a.municipio, a.provincia])

    col_widths = [35, 15, 40, 10, 20, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def exportar_informe_comparacion_iva(anio, autoescuela):
    from facturacion.utils.iva_comparator import comparar_iva_anual

    informe = comparar_iva_anual(anio, autoescuela)
    wb = Workbook()

    ws = wb.active
    ws.title = 'RESUMEN'
    ws.append(['Trimestre', 'Facturas Correctas', 'IVA Correcto', 'Tasas (exentas)',
               'IVA Declarado (incorrecto)', 'Diferencia IVA'])

    for r in informe['resultados']:
        ws.append([
            f'T{r["trimestre"]}',
            r['correcto']['num_facturas'],
            float(r['correcto']['iva']),
            float(r['correcto']['tasas']),
            float(r['incorrecto']['iva']),
            float(r['diferencia_iva']),
        ])

    ws.append([
        'TOTAL ANUAL',
        sum(r['correcto']['num_facturas'] for r in informe['resultados']),
        float(informe['total_iva_correcto']),
        float(informe['total_tasas']),
        float(informe['total_iva_incorrecto']),
        float(informe['diferencia_total']),
    ])

    for col in range(3, 7):
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = '#,##0.00'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
