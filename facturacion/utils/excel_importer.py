"""
Importador de archivos Excel a la base de datos.
"""
from openpyxl import load_workbook
from datetime import datetime
import re


def normalize_dni(dni):
    if not dni:
        return ''
    return str(dni).upper().replace(' ', '').replace('-', '').strip()


def importar_trimestre(file_obj, trimestre, anio, autoescuela):
    """
    Importa facturas desde un archivo Trimestre-X.xlsx.
    Headers: CURSO, Nº FACTURA, FECHA, NOMBRE Y APELLIDOS, DNI,
             BASE IMPONIBLE, IVA, TASAS, TOTAL, DIRECCION, CP, MUNICIPIO, PROVINCIA
    """
    from facturacion.models import Factura, Alumno

    wb = load_workbook(file_obj, read_only=True)
    ws = wb.active
    count = 0

    for row in range(2, ws.max_row + 1):
        nfactura = ws.cell(row=row, column=2).value
        if nfactura is None:
            continue

        curso = str(ws.cell(row=row, column=1).value or '').strip()
        nfactura_str = str(nfactura).strip()
        fecha_raw = ws.cell(row=row, column=3).value
        nombre = str(ws.cell(row=row, column=4).value or '').strip()
        dni = str(ws.cell(row=row, column=5).value or '').strip()
        base = ws.cell(row=row, column=6).value or 0
        iva = ws.cell(row=row, column=7).value or 0
        tasas = ws.cell(row=row, column=8).value or 0
        total = ws.cell(row=row, column=9).value or 0
        direccion = str(ws.cell(row=row, column=10).value or '').strip()
        cp = str(ws.cell(row=row, column=11).value or '').strip()
        municipio = str(ws.cell(row=row, column=12).value or '').strip()
        provincia = str(ws.cell(row=row, column=13).value or '').strip()

        # Parsear fecha
        if isinstance(fecha_raw, datetime):
            fecha = fecha_raw.date()
        elif fecha_raw:
            try:
                fecha = datetime.strptime(str(fecha_raw).strip(), '%d/%m/%Y').date()
            except ValueError:
                try:
                    fecha = datetime.fromisoformat(str(fecha_raw).strip()).date()
                except ValueError:
                    fecha = datetime(anio, trimestre * 3, 1).date()
        else:
            fecha = datetime(anio, trimestre * 3, 1).date()

        # Normalizar número de factura
        nfactura_str = normalizar_num_factura(nfactura_str, anio)

        # Buscar o crear alumno (dirección opcional)
        alumno = None
        if dni:
            alumno, created = Alumno.objects.get_or_create(
                dni=dni,
                autoescuela=autoescuela,
                defaults={
                    'nombre': nombre,
                    'direccion': direccion,
                    'codigo_postal': cp,
                    'municipio': municipio,
                    'provincia': provincia,
                }
            )
            if not created and direccion:
                alumno.nombre = nombre
                alumno.direccion = direccion
                alumno.codigo_postal = cp
                alumno.municipio = municipio
                alumno.provincia = provincia
                alumno.save()

        # Crear factura (skip si ya existe en esta autoescuela)
        if not Factura.objects.filter(numero_factura=nfactura_str, autoescuela=autoescuela).exists():
            Factura.objects.create(
                curso=curso,
                numero_factura=nfactura_str,
                fecha=fecha,
                alumno=alumno,
                nombre_factura=nombre,
                dni_factura=dni,
                direccion_factura=direccion,
                cp_factura=cp,
                municipio_factura=municipio,
                provincia_factura=provincia,
                base_imponible=float(base),
                iva=float(iva),
                tasas=float(tasas),
                total=float(total),
                trimestre=trimestre,
                anio=anio,
                autoescuela=autoescuela,
            )
            count += 1

    wb.close()
    return count


def normalizar_num_factura(num_str, anio_default):
    """Normaliza número de factura al formato YYYY/NNNN."""
    num_str = str(num_str).strip()
    match = re.match(r'(\d{4})/(\d+)', num_str)
    if match:
        year = match.group(1)
        numero = match.group(2).zfill(4)
        return f"{year}/{numero}"
    # Solo número
    try:
        n = int(num_str)
        return f"{anio_default}/{n:04d}"
    except ValueError:
        return num_str


def importar_listado(file_obj, trimestre, anio, autoescuela):
    """
    Importa datos de LISTADO FACT X TRI.xlsx (datos históricos con IVA incorrecto).
    """
    from facturacion.models import ListadoHistorico
    import pandas as pd

    if trimestre in [1, 2]:
        df = pd.read_excel(file_obj, header=1)
        cols = df.columns.tolist()
        factura_col = [c for c in cols if 'FRA' in str(c).upper()][0]
        iva_col = [c for c in cols if 'IVA' in str(c).upper()][0]
        neto_col = [c for c in cols if 'NETO' in str(c).upper()][0]
        total_col = [c for c in cols if 'TOTAL' in str(c).upper()][0]
    else:
        df = pd.read_excel(file_obj, header=None)
        df.columns = df.iloc[0]
        df = df[1:]
        factura_col = 'Nº FRA.'
        neto_col = 'NETO'
        iva_col = 'IVA 21%'
        total_col = 'TOTAL'

    df = df.dropna(subset=[factura_col])

    # Convertir columnas numéricas y reemplazar NaN por 0
    for col in [neto_col, iva_col, total_col]:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    count = 0

    for _, row in df.iterrows():
        nf = normalizar_num_factura(str(row[factura_col]), anio)
        neto = float(row[neto_col])
        iva_val = float(row[iva_col])
        total_val = float(row[total_col])

        ListadoHistorico.objects.update_or_create(
            numero_factura=nf,
            trimestre=trimestre,
            anio=anio,
            autoescuela=autoescuela,
            defaults={
                'neto': neto,
                'iva_incorrecto': iva_val,
                'total': total_val,
            }
        )
        count += 1

    return count
